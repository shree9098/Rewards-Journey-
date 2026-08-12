"""Create a seasonal SKU offer sheet from a warehouse Excel workbook."""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


@dataclass(frozen=True)
class Offer:
    name: str
    discount: float
    max_price: float
    sku_count: int


# Change or extend these defaults whenever a new campaign is introduced.
SEASONAL_OFFERS = {
    "republic-day": Offer("Republic Day", 0.15, 1000, 12),
    "holi": Offer("Holi", 0.10, 1000, 12),
    "summer": Offer("Summer", 0.15, 1500, 12),
    "independence-day": Offer("Independence Day", 0.15, 1000, 12),
    "raksha-bandhan": Offer("Raksha Bandhan", 0.15, 1500, 12),
    "diwali": Offer("Diwali", 0.20, 2000, 12),
    "christmas": Offer("Christmas", 0.20, 2000, 12),
    "new-year": Offer("New Year", 0.20, 2000, 12),
}

HEADER_ALIASES = {
    "brand": {"BRAND"},
    "product": {"PRODUCT NAME", "PRODUCT", "ITEM NAME", "ITEM"},
    "cost": {"B2B", "B2B COST", "COST"},
    "price": {"RP PRICING", "RP RPICING", "RETAIL PRICE", "PRICE", "MRP"},
}


def normalize(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().strip("'\"").upper())


def safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "-", name).strip()
    return (cleaned or "Seasonal Offer")[:31]


def find_columns(sheet) -> dict[str, int]:
    normalized = {normalize(cell.value): cell.column for cell in sheet[1]}
    found: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                found[field] = normalized[alias]
                break
    missing = sorted(set(HEADER_ALIASES) - set(found))
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")
    return found


def choose_products(sheet, columns: dict[str, int], offer: Offer) -> list[dict]:
    eligible: list[dict] = []
    for row in range(2, sheet.max_row + 1):
        try:
            price = float(sheet.cell(row, columns["price"]).value)
            cost = float(sheet.cell(row, columns["cost"]).value)
        except (TypeError, ValueError):
            continue
        product = sheet.cell(row, columns["product"]).value
        if product and 0 < price <= offer.max_price:
            eligible.append(
                {
                    "brand": sheet.cell(row, columns["brand"]).value,
                    "product": product,
                    "cost": cost,
                    "price": price,
                    "row": row,
                }
            )

    if len(eligible) < offer.sku_count:
        raise ValueError(
            f"Only {len(eligible)} eligible products are priced at or below "
            f"Rs {offer.max_price:,.2f}; {offer.sku_count} were requested"
        )

    # Spread selection across brands, then favor higher-value products.
    eligible.sort(key=lambda item: (-item["price"], normalize(item["brand"]), item["row"]))
    selected: list[dict] = []
    remaining = eligible[:]
    used_brands: set[str] = set()
    while remaining and len(selected) < offer.sku_count:
        different_brand = next(
            (item for item in remaining if normalize(item["brand"]) not in used_brands), None
        )
        item = different_brand or remaining[0]
        selected.append(item)
        used_brands.add(normalize(item["brand"]))
        remaining.remove(item)
    return selected


def style_sheet(sheet) -> None:
    navy, saffron, green = "17365D", "FF9933", "138808"
    thin = Side(style="thin", color="B7C9D6")
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=saffron))
    for row in range(2, sheet.max_row + 1):
        for cell in sheet[row]:
            cell.fill = PatternFill("solid", fgColor="FFFFFF" if row % 2 == 0 else "EAF2F8")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=cell.column == 3)
        sheet.cell(row, 6).number_format = "0%"
        for column in (4, 5, 7, 8):
            sheet.cell(row, column).number_format = '[$Rs-en-IN]#,##0.00'
    for column, width in {
        "A": 10, "B": 18, "C": 58, "D": 15, "E": 16, "F": 13, "G": 18, "H": 17
    }.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = green


def populate_sheet(sheet, products: list[dict], offer: Offer) -> None:
    sheet.append(
        ["SKU No.", "Brand", "Product Name", "B2B Cost", "Original Price",
         "Discount %", "Price Difference", "Offer Price"]
    )
    for sku, item in enumerate(products, 1):
        row = sheet.max_row + 1
        sheet.append(
            [sku, item["brand"], item["product"], item["cost"], item["price"],
             offer.discount, f"=ROUND(E{row}*F{row},2)", f"=ROUND(E{row}-G{row},2)"]
        )
    style_sheet(sheet)


def generate_offer(
    source: Path, offer: Offer, source_sheet: str | None = None,
    output: Path | None = None, separate_file: Path | None = None,
) -> tuple[Path, Path | None]:
    if not source.is_file():
        raise FileNotFoundError(f"Workbook not found: {source}")
    workbook = load_workbook(source)
    source_ws = workbook[source_sheet] if source_sheet else workbook.worksheets[0]
    products = choose_products(source_ws, find_columns(source_ws), offer)
    sheet_name = safe_sheet_name(f"{offer.name} SKU Sheet")
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    populate_sheet(workbook.create_sheet(sheet_name), products, offer)
    destination = output or source
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)

    if separate_file:
        separate_file.parent.mkdir(parents=True, exist_ok=True)
        standalone = Workbook()
        standalone.remove(standalone.active)
        # Rebuild so formulas, formatting, and widths remain portable.
        populate_sheet(standalone.create_sheet(sheet_name), products, offer)
        standalone.save(separate_file)
    return destination, separate_file


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source", nargs="?", type=Path, default=Path("Warehouse_Stock.xlsx"))
    parser.add_argument("--season", choices=sorted(SEASONAL_OFFERS), default="independence-day")
    parser.add_argument("--name", help="custom offer name; overrides the preset name")
    parser.add_argument("--discount", type=float, help="discount percentage, for example 15")
    parser.add_argument("--max-price", type=float, help="maximum original price")
    parser.add_argument("--sku-count", type=int, help="number of SKUs to select")
    parser.add_argument("--source-sheet", help="source worksheet; first sheet is used by default")
    parser.add_argument("--output", type=Path, help="output workbook; source is updated by default")
    parser.add_argument("--separate-file", type=Path, help="also create a standalone offer workbook")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    preset = SEASONAL_OFFERS[args.season]
    offer = Offer(
        args.name or preset.name,
        (args.discount / 100) if args.discount is not None else preset.discount,
        args.max_price if args.max_price is not None else preset.max_price,
        args.sku_count if args.sku_count is not None else preset.sku_count,
    )
    if not 0 <= offer.discount <= 1 or offer.max_price <= 0 or offer.sku_count <= 0:
        print("Error: discount must be 0-100; max price and SKU count must be positive", file=sys.stderr)
        return 1
    try:
        destination, separate = generate_offer(
            args.source, offer, args.source_sheet, args.output, args.separate_file
        )
    except (FileNotFoundError, KeyError, ValueError, OSError) as error:
        print(f"Error: {error}", file=sys.stderr)
        return 1
    print(f"Created {safe_sheet_name(offer.name + ' SKU Sheet')} in {destination.resolve()}")
    if separate:
        print(f"Standalone file: {separate.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
